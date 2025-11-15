"""
Servicio para manejar importación y exportación de datos en Excel
"""
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.models.database import DatabaseManager

logger = logging.getLogger(__name__)


class ExcelService:
    """Servicio para manejar operaciones con archivos Excel"""
    
    def __init__(self, db_manager: DatabaseManager):
        """
        Inicializa el servicio de Excel
        
        Args:
            db_manager: Instancia del gestor de base de datos
        """
        self.db_manager = db_manager
    
    def exportar_a_excel(self, ruta_archivo: Path, formato_por_defecto: Optional[str] = None) -> Tuple[bool, str]:
        """
        Exporta todos los datos de la base de datos a un archivo Excel
        
        Args:
            ruta_archivo: Ruta donde guardar el archivo Excel
            formato_por_defecto: Formato a usar si algún registro no tiene formato (opcional)
            
        Returns:
            Tupla (éxito, mensaje)
        """
        try:
            # Obtener todos los códigos de la base de datos
            codigos = self.db_manager.obtener_todos_codigos()
            
            if not codigos:
                return False, "No hay datos en la base de datos para exportar"
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Códigos de Barras"
            
            # Obtener estructura dinámica de la base de datos
            # Las columnas son: id, codigo_barras, id_unico, fecha_creacion, 
            # nombre_empleado, descripcion, formato, nombre_archivo
            # Nota: nombre_archivo se excluye de la exportación de códigos de barras
            columnas = [
                "ID",
                "Código de Barras",
                "ID Único",
                "Fecha de Creación",
                "Nombre del Empleado",
                "Código de Empleado",
                "Formato"
            ]
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir encabezados
            for col_idx, columna in enumerate(columnas, start=1):
                cell = ws.cell(row=1, column=col_idx, value=columna)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Escribir datos (excluyendo nombre_archivo que es el último campo)
            for row_idx, codigo in enumerate(codigos, start=2):
                # codigo es una tupla: (id, codigo_barras, id_unico, fecha_creacion, 
                # nombre_empleado, descripcion, formato, nombre_archivo)
                # Solo escribimos los primeros 7 campos (excluyendo nombre_archivo)
                for col_idx in range(len(columnas)):
                    valor = codigo[col_idx] if col_idx < len(codigo) else ""
                    # Si es la columna de formato (índice 6) y está vacío, usar formato por defecto
                    if col_idx == 6 and (not valor or valor == ""):
                        valor = formato_por_defecto or "Code128"
                    ws.cell(row=row_idx, column=col_idx + 1, value=valor)
            
            # Ajustar ancho de columnas
            for col_idx in range(1, len(columnas) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 20
            
            # Guardar archivo
            wb.save(str(ruta_archivo))
            logger.info(f"Excel exportado exitosamente: {ruta_archivo}")
            return True, f"Se exportaron {len(codigos)} registros exitosamente"
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {e}", exc_info=True)
            return False, f"Error al exportar: {str(e)}"
    
    def generar_excel_ejemplo(self, ruta_archivo: Path, formato_por_defecto: Optional[str] = None) -> Tuple[bool, str]:
        """
        Genera un archivo Excel de ejemplo con el formato esperado
        
        Args:
            ruta_archivo: Ruta donde guardar el archivo Excel de ejemplo
            formato_por_defecto: Formato a usar en todas las filas de ejemplo (opcional)
            
        Returns:
            Tupla (éxito, mensaje)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos de Empleados"
            
            # Columnas requeridas y opcionales
            columnas = [
                "Nombre del Empleado",  # Obligatorio
                "Código de Empleado",   # Obligatorio
                "Formato (opcional)"    # Opcional (por defecto Code128)
            ]
            
            # Usar formato por defecto o Code128
            formato_a_usar = formato_por_defecto or "Code128"
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir encabezados
            for col_idx, columna in enumerate(columnas, start=1):
                cell = ws.cell(row=1, column=col_idx, value=columna)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Datos de ejemplo (todos con el formato seleccionado)
            datos_ejemplo = [
                ("Juan Pérez", "EMP001", formato_a_usar),
                ("María García", "EMP002", formato_a_usar),
                ("Carlos López", "EMP003", formato_a_usar),
            ]
            
            # Escribir datos de ejemplo
            for row_idx, datos in enumerate(datos_ejemplo, start=2):
                for col_idx, valor in enumerate(datos, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=valor)
            
            # Ajustar ancho de columnas
            for col_idx in range(1, len(columnas) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 25
            
            # Guardar archivo
            wb.save(str(ruta_archivo))
            logger.info(f"Excel de ejemplo generado: {ruta_archivo}")
            return True, "Archivo de ejemplo generado exitosamente"
            
        except Exception as e:
            logger.error(f"Error al generar Excel de ejemplo: {e}", exc_info=True)
            return False, f"Error al generar ejemplo: {str(e)}"
    
    def importar_desde_excel(
        self, 
        ruta_archivo: Path,
        callback_progreso: Optional[callable] = None,
        formato_por_defecto: Optional[str] = None
    ) -> Tuple[bool, Dict[str, int], List[str]]:
        """
        Importa datos desde un archivo Excel y genera códigos de barras
        
        Args:
            ruta_archivo: Ruta al archivo Excel a importar
            callback_progreso: Función callback para reportar progreso
                              Recibe (actual, total, mensaje)
            formato_por_defecto: Formato a usar cuando no se especifica en el Excel (opcional)
            
        Returns:
            Tupla (éxito, estadísticas, errores)
            estadísticas: {'exitosos': int, 'errores': int, 'duplicados': int, 'validacion_fallida': int}
            errores: Lista de mensajes de error
        """
        try:
            if not ruta_archivo.exists():
                return False, {}, ["El archivo Excel no existe"]
            
            # Cargar workbook
            wb = openpyxl.load_workbook(str(ruta_archivo), data_only=True)
            ws = wb.active
            
            # Leer encabezados
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else "")
            
            # Buscar índices de columnas requeridas
            try:
                idx_nombre = headers.index("Nombre del Empleado")
                idx_codigo_empleado = headers.index("Código de Empleado")
                # Buscar "Formato (opcional)" o "Formato" para compatibilidad
                idx_formato = None
                if "Formato (opcional)" in headers:
                    idx_formato = headers.index("Formato (opcional)")
                elif "Formato" in headers:
                    idx_formato = headers.index("Formato")
            except ValueError as e:
                return False, {}, [f"Columna requerida no encontrada: {str(e)}"]
            
            # Estadísticas
            estadisticas = {
                'exitosos': 0,
                'errores': 0,
                'duplicados': 0,
                'validacion_fallida': 0,
                'total': 0
            }
            errores = []
            
            # Procesar filas (empezar desde la fila 2, saltando encabezados)
            total_filas = ws.max_row - 1  # Excluir encabezado
            estadisticas['total'] = total_filas
            
            for row_idx in range(2, ws.max_row + 1):
                if callback_progreso:
                    callback_progreso(
                        row_idx - 1, 
                        total_filas, 
                        f"Procesando fila {row_idx - 1} de {total_filas}..."
                    )
                
                # Leer datos de la fila
                nombre_empleado = ws.cell(row=row_idx, column=idx_nombre + 1).value
                codigo_empleado = ws.cell(row=row_idx, column=idx_codigo_empleado + 1).value
                formato = ws.cell(row=row_idx, column=idx_formato + 1).value if idx_formato is not None else None
                
                # Validar datos obligatorios
                if not nombre_empleado or not codigo_empleado:
                    estadisticas['errores'] += 1
                    errores.append(f"Fila {row_idx}: Faltan datos obligatorios (Nombre o Código de Empleado)")
                    continue
                
                # Limpiar valores
                nombre_empleado = str(nombre_empleado).strip()
                codigo_empleado = str(codigo_empleado).strip()
                # Usar formato del Excel si existe, sino el formato por defecto, sino Code128
                formato = str(formato).strip() if formato else (formato_por_defecto or "Code128")
                
                # Validar formato
                formatos_validos = ["Code128", "EAN13", "EAN8", "Code39"]
                if formato not in formatos_validos:
                    formato = "Code128"  # Usar por defecto
                
                # Verificar si el código de empleado ya existe (buscar por descripcion que es el código de empleado)
                # La búsqueda puede retornar múltiples resultados, pero buscamos por descripción exacta
                todos_codigos = self.db_manager.obtener_todos_codigos()
                codigo_existente = None
                for codigo in todos_codigos:
                    # codigo es: (id, codigo_barras, id_unico, fecha_creacion, nombre_empleado, descripcion, formato, nombre_archivo)
                    if len(codigo) >= 6 and codigo[5] == codigo_empleado:  # descripcion es índice 5
                        codigo_existente = codigo
                        break
                
                if codigo_existente:
                    # Si existe, verificar si tiene código de barras válido
                    id_db, codigo_barras, id_unico, fecha, nombre, descripcion, formato_existente, nombre_archivo = codigo_existente
                    
                    # Verificar si el código de barras es válido
                    from config.settings import IMAGES_DIR
                    from src.services.barcode_service import BarcodeService
                    barcode_service = BarcodeService()
                    
                    if nombre_archivo:
                        ruta_imagen = IMAGES_DIR / nombre_archivo
                        if ruta_imagen.exists():
                            valido, mensaje = barcode_service.validar_codigo_barras(ruta_imagen, id_unico)
                            if not valido:
                                estadisticas['validacion_fallida'] += 1
                                errores.append(
                                    f"Fila {row_idx} ({nombre_empleado}): "
                                    f"El código de barras existente no es válido. "
                                    f"¿Desea regenerarlo? (Se requiere confirmación manual)"
                                )
                                continue
                    
                    estadisticas['duplicados'] += 1
                    errores.append(
                        f"Fila {row_idx} ({nombre_empleado}): "
                        f"El código de empleado '{codigo_empleado}' ya existe en la base de datos"
                    )
                    continue
                
                # Datos válidos para procesar
                estadisticas['exitosos'] += 1
            
            return True, estadisticas, errores
            
        except Exception as e:
            logger.error(f"Error al importar desde Excel: {e}", exc_info=True)
            return False, {}, [f"Error al importar: {str(e)}"]

