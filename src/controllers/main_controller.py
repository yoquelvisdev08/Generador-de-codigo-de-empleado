"""
Controlador principal - Coordina la lógica de presentación
"""
import os
from pathlib import Path
from typing import Optional
from datetime import datetime
from PyQt6.QtWidgets import QMessageBox, QFileDialog

from config.settings import IMAGES_DIR
from src.models.database import DatabaseManager
from src.models.barcode_model import BarcodeModel
from src.services.barcode_service import BarcodeService
from src.services.export_service import ExportService
from src.services.excel_service import ExcelService
from src.views.main_window import MainWindow
from src.views.widgets.progress_dialog import ProgressDialog
from src.utils.file_utils import obtener_ruta_imagen
from src.utils.auth_utils import solicitar_password
from src.utils.id_generator import IDGenerator
from src.controllers.carnet_controller import CarnetController


class MainController:
    """Controlador principal que coordina modelos, servicios y vistas"""
    
    def __init__(self, usuario: str = "admin", rol: str = "admin"):
        """
        Inicializa el controlador y sus dependencias
        
        Args:
            usuario: Nombre del usuario autenticado
            rol: Rol del usuario ('admin' o 'user')
        """
        self.usuario = usuario
        self.rol = rol
        
        self.db_manager = DatabaseManager()
        self.barcode_service = BarcodeService()
        self.export_service = ExportService()
        self.excel_service = ExcelService(self.db_manager)
        
        # Configurar formatos disponibles en el panel de generación
        formatos = self.barcode_service.obtener_formatos_disponibles()
        self.main_window = MainWindow(formatos_disponibles=formatos)
        
        # Configurar visibilidad de botones según el rol
        self.main_window.list_panel.configurar_permisos(es_admin=(rol == "admin"))
        
        # Inicializar controlador de carnet (se inicializa cuando se muestra la vista)
        self.carnet_controller = None
        
        self._conectar_senales()
        self._cargar_datos_iniciales()
    
    def _conectar_senales(self):
        """Conecta las señales de los widgets con los métodos del controlador"""
        # Navbar - navegación principal (menú Tools)
        self.main_window.accion_codigo_barras.triggered.connect(self.mostrar_vista_generacion)
        self.main_window.accion_crear_carnet.triggered.connect(self.mostrar_vista_carnet)
        
        # Panel de generación
        self.main_window.generation_panel.boton_generar.clicked.connect(self.generar_codigo)
        # Conectar señales para actualización en tiempo real del ID
        self.main_window.generation_panel.conectar_senales_actualizacion(self.actualizar_id_preview)
        
        # Panel de carnet - El botón volver fue eliminado, la navegación se hace desde el navbar
        
        # Panel de listado
        self.main_window.list_panel.campo_busqueda.textChanged.connect(self.buscar_codigos)
        self.main_window.list_panel.boton_refrescar.clicked.connect(self.cargar_codigos)
        # Mostrar imagen automáticamente al seleccionar una fila
        self.main_window.list_panel.tabla_codigos.itemSelectionChanged.connect(self.mostrar_imagen_seleccionada)
        self.main_window.list_panel.tabla_codigos.doubleClicked.connect(self.mostrar_detalle_codigo)
        self.main_window.list_panel.boton_exportar.clicked.connect(self.exportar_seleccionados)
        self.main_window.list_panel.boton_exportar_todos.clicked.connect(self.exportar_todos_zip)
        self.main_window.list_panel.boton_backup.clicked.connect(self.backup_base_datos)
        self.main_window.list_panel.boton_eliminar.clicked.connect(self.eliminar_codigo)
        self.main_window.list_panel.boton_limpiar.clicked.connect(self.limpiar_base_datos)
        self.main_window.list_panel.boton_limpiar_imagenes.clicked.connect(self.limpiar_imagenes_huerfanas)
        # Botones de Excel
        self.main_window.list_panel.boton_exportar_excel.clicked.connect(self.exportar_data_excel)
        self.main_window.list_panel.boton_importar_excel.clicked.connect(self.importar_data_excel)
        self.main_window.list_panel.boton_descargar_ejemplo_excel.clicked.connect(self.descargar_ejemplo_excel)
    
    def _cargar_datos_iniciales(self):
        """Carga los datos iniciales en la interfaz"""
        self.cargar_codigos()
        self.actualizar_id_preview()
        self.actualizar_estadisticas()
    
    def generar_codigo(self):
        """Genera un nuevo código de barras"""
        datos = self.main_window.generation_panel.obtener_datos()
        nombre_empleado = datos['nombre_empleado']
        formato = datos['formato']
        codigo_empleado = datos['descripcion']  # Mantener nombre interno como 'descripcion' para compatibilidad con BD
        opciones_id = self.main_window.generation_panel.obtener_opciones_id()
        
        if not nombre_empleado:
            QMessageBox.warning(
                self.main_window, "Advertencia",
                "Por favor ingrese el nombre del empleado"
            )
            return
        
        if not codigo_empleado:
            QMessageBox.warning(
                self.main_window, "Advertencia",
                "Por favor ingrese el código de empleado (campo obligatorio)"
            )
            return
        
        try:
            # Generar ID personalizado según las opciones seleccionadas
            id_unico_generado = IDGenerator.generar_id_personalizado(
                tipo=opciones_id['tipo_caracteres'],
                longitud=opciones_id['cantidad_caracteres'],
                incluir_nombre=opciones_id['incluir_nombre'],
                nombre_empleado=nombre_empleado if opciones_id['incluir_nombre'] else None,
                texto_personalizado=opciones_id.get('texto_personalizado'),
                verificar_duplicado=self.db_manager.verificar_codigo_existe
            )
            
            codigo_barras, id_unico_archivo, ruta_imagen = self.barcode_service.generar_codigo_barras(
                id_unico_generado, formato, id_unico_generado, nombre_empleado
            )
            
            valido, mensaje_error = self.barcode_service.validar_codigo_barras(
                ruta_imagen, id_unico_generado
            )
            
            if not valido:
                if ruta_imagen.exists():
                    ruta_imagen.unlink()
                QMessageBox.critical(
                    self.main_window, "Error de Validación",
                    f"El código de barras generado no es válido:\n{mensaje_error}\n\n"
                    f"El código no se ha guardado. Por favor intente generar nuevamente."
                )
                return
            
            nombre_archivo = ruta_imagen.name
            
            exito = self.db_manager.insertar_codigo(
                codigo_barras, id_unico_generado, formato, nombre_empleado, codigo_empleado, nombre_archivo
            )
            
            if not exito:
                if ruta_imagen.exists():
                    ruta_imagen.unlink()
                QMessageBox.warning(
                    self.main_window, "Error",
                    "No se pudo guardar el código en la base de datos.\n"
                    "Puede que el ID único ya exista."
                )
                return
            
            self.main_window.generation_panel.mostrar_vista_previa(str(ruta_imagen))
            self.cargar_codigos()
            self.actualizar_estadisticas()
            self.actualizar_id_preview()
            self.main_window.generation_panel.limpiar_formulario()
            
            # Actualizar tabla de empleados en la vista de carnet si está activa
            if self.carnet_controller is not None:
                self.carnet_controller.refrescar_empleados()
            
            QMessageBox.information(
                self.main_window, "Éxito",
                f"Código de barras generado y validado exitosamente.\n"
                f"ID Único: {id_unico_generado}\n"
                f"Empleado: {nombre_empleado}\n"
                f"Validación: El código escaneado coincide con el ID generado"
            )
        except Exception as e:
            QMessageBox.critical(
                self.main_window, "Error",
                f"Error al generar el código de barras:\n{str(e)}"
            )
    
    def cargar_codigos(self):
        """Carga todos los códigos en la tabla"""
        termino = self.main_window.list_panel.obtener_termino_busqueda()
        
        if termino:
            codigos = self.db_manager.buscar_codigo(termino)
        else:
            codigos = self.db_manager.obtener_todos_codigos()
        
        self.main_window.list_panel.cargar_codigos(codigos)
    
    def buscar_codigos(self):
        """Busca códigos según el término de búsqueda"""
        self.cargar_codigos()
    
    def mostrar_imagen_seleccionada(self):
        """Muestra automáticamente la imagen del código seleccionado en la tabla"""
        resultado = self.main_window.list_panel.obtener_fila_seleccionada()
        if resultado is None:
            # Si no hay selección, limpiar la vista previa
            return
        
        id_db, codigo_barras, id_unico, formato, nombre_archivo = resultado
        ruta_imagen = IMAGES_DIR / nombre_archivo
        
        if ruta_imagen.exists():
            self.main_window.generation_panel.mostrar_vista_previa(str(ruta_imagen))
        else:
            # Si la imagen no existe, mostrar mensaje en la vista previa
            self.main_window.generation_panel.label_vista_previa.setText(
                f"Imagen no encontrada:\n{nombre_archivo}"
            )
    
    def mostrar_detalle_codigo(self):
        """Muestra el detalle del código seleccionado (doble clic)"""
        self.mostrar_imagen_seleccionada()
    
    def exportar_seleccionados(self):
        """Exporta los códigos seleccionados"""
        filas = self.main_window.list_panel.obtener_filas_seleccionadas()
        
        if not filas:
            QMessageBox.warning(
                self.main_window, "Advertencia",
                "Por favor seleccione al menos un código de la tabla"
            )
            return
        
        directorio_destino = QFileDialog.getExistingDirectory(
            self.main_window, "Seleccionar carpeta para exportar"
        )
        
        if not directorio_destino:
            return
        
        exportados, errores = self.export_service.exportar_seleccionados(
            filas, Path(directorio_destino)
        )
        
        mensaje = f"Exportación completada:\n{exportados} archivo(s) exportado(s)"
        if errores > 0:
            mensaje += f"\n{errores} error(es)"
        
        QMessageBox.information(self.main_window, "Exportación", mensaje)
    
    def exportar_todos_zip(self):
        """Exporta todos los códigos a un archivo ZIP"""
        codigos = self.db_manager.obtener_todos_codigos()
        
        if not codigos:
            QMessageBox.warning(
                self.main_window, "Advertencia",
                "No hay códigos para exportar"
            )
            return
        
        nombre_zip = self.export_service.generar_nombre_zip()
        ruta_zip, _ = QFileDialog.getSaveFileName(
            self.main_window, "Guardar archivo ZIP",
            nombre_zip,
            "Archivos ZIP (*.zip);;Todos los archivos (*)"
        )
        
        if not ruta_zip:
            return
        
        agregados, errores = self.export_service.exportar_todos_zip(
            codigos, Path(ruta_zip)
        )
        
        mensaje = f"ZIP creado exitosamente:\n{agregados} archivo(s) incluido(s)"
        if errores > 0:
            mensaje += f"\n{errores} archivo(s) no encontrado(s)"
        
        QMessageBox.information(self.main_window, "Éxito", mensaje)
    
    def backup_base_datos(self):
        """Crea un backup de la base de datos"""
        nombre_backup = self.export_service.generar_nombre_backup()
        ruta_backup, _ = QFileDialog.getSaveFileName(
            self.main_window, "Guardar Backup",
            nombre_backup,
            "Archivos de Base de Datos (*.db);;Todos los archivos (*)"
        )
        
        if not ruta_backup:
            return
        
        if self.export_service.crear_backup_base_datos(Path(ruta_backup)):
            QMessageBox.information(
                self.main_window, "Éxito",
                f"Backup creado exitosamente:\n{ruta_backup}"
            )
        else:
            QMessageBox.critical(
                self.main_window, "Error",
                "Error al crear el backup"
            )
    
    def eliminar_codigo(self):
        """Elimina un código de la base de datos"""
        resultado = self.main_window.list_panel.obtener_fila_seleccionada()
        if resultado is None:
            QMessageBox.warning(
                self.main_window, "Advertencia",
                "Por favor seleccione un código de la tabla"
            )
            return
        
        id_db, codigo_barras, id_unico, formato, nombre_archivo = resultado
        
        # Solicitar contraseña antes de eliminar
        if not solicitar_password(self.main_window, "eliminar este código"):
            return
        
        respuesta = QMessageBox.question(
            self.main_window, "Confirmar Eliminación",
            f"¿Está seguro de que desea eliminar este código?\n"
            f"ID Único: {id_unico}\n\n"
            "Nota: La imagen también se eliminará del disco.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta == QMessageBox.StandardButton.Yes:
            if self.db_manager.eliminar_codigo(id_db, eliminar_imagen=True):
                QMessageBox.information(
                    self.main_window, "Éxito",
                    "Código e imagen eliminados correctamente.\n\n"
                    "Nota: Se creó un backup automático antes de la eliminación."
                )
                self.cargar_codigos()
                self.actualizar_estadisticas()
            else:
                QMessageBox.warning(
                    self.main_window, "Error",
                    "No se pudo eliminar el código"
                )
    
    def limpiar_base_datos(self):
        """Limpia toda la base de datos"""
        # Solicitar contraseña antes de limpiar la base de datos
        if not solicitar_password(self.main_window, "limpiar toda la base de datos"):
            return
        
        respuesta = QMessageBox.question(
            self.main_window, "Confirmar Limpieza",
            "¿Está seguro de que desea eliminar TODOS los códigos de la base de datos?\n\n"
            "Esta acción NO se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if respuesta == QMessageBox.StandardButton.Yes:
            if self.db_manager.limpiar_base_datos(eliminar_imagenes=True):
                QMessageBox.information(
                    self.main_window, "Éxito",
                    "Base de datos e imágenes limpiadas exitosamente.\n\n"
                    "Nota: Se creó un backup automático antes de la limpieza."
                )
                self.cargar_codigos()
                self.actualizar_estadisticas()
                self.actualizar_id_preview()
            else:
                QMessageBox.warning(
                    self.main_window, "Error",
                    "No se pudo limpiar la base de datos.\n\n"
                    "Verifique que se haya creado el backup automático."
                )
    
    def limpiar_imagenes_huerfanas(self):
        """Limpia imágenes huérfanas (imágenes sin registro en la BD)"""
        respuesta = QMessageBox.question(
            self.main_window, "Confirmar Limpieza",
            "¿Desea eliminar las imágenes que no tienen registro en la base de datos?\n\n"
            "Esto eliminará solo las imágenes huérfanas, no afectará los códigos existentes.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if respuesta == QMessageBox.StandardButton.Yes:
            eliminadas, errores = self.db_manager.limpiar_imagenes_huerfanas()
            
            mensaje = f"Limpieza completada:\n{eliminadas} imagen(es) huérfana(s) eliminada(s)"
            if errores > 0:
                mensaje += f"\n{errores} error(es) al eliminar"
            
            QMessageBox.information(self.main_window, "Limpieza de Imágenes", mensaje)
    
    def actualizar_id_preview(self):
        """Actualiza la vista previa del ID que se generará en tiempo real"""
        try:
            opciones = self.main_window.generation_panel.obtener_opciones_id()
            nombre_empleado = opciones['nombre_empleado']
            incluir_nombre = opciones['incluir_nombre']
            
            # Si se debe incluir el nombre pero no hay nombre, mostrar mensaje
            if incluir_nombre and not nombre_empleado:
                self.main_window.generation_panel.actualizar_id_preview(
                    "Ingrese el nombre del empleado para ver el ID"
                )
                return
            
            # Generar un ID de ejemplo (sin verificar duplicados para la vista previa)
            id_ejemplo = IDGenerator.generar_id_personalizado(
                tipo=opciones['tipo_caracteres'],
                longitud=opciones['cantidad_caracteres'],
                incluir_nombre=incluir_nombre,
                nombre_empleado=nombre_empleado if incluir_nombre else None,
                texto_personalizado=opciones.get('texto_personalizado'),
                verificar_duplicado=None  # No verificar en la vista previa
            )
            
            self.main_window.generation_panel.actualizar_id_preview(id_ejemplo)
        except Exception:
            # Si hay error, usar el método tradicional
            siguiente_id = self.db_manager.obtener_siguiente_id_secuencial()
            self.main_window.generation_panel.actualizar_id_preview(siguiente_id)
    
    def actualizar_estadisticas(self):
        """Actualiza las estadísticas en la barra de estado"""
        estadisticas = self.db_manager.obtener_estadisticas()
        mensaje = (
            f"Total de códigos: {estadisticas['total_codigos']} | "
            f"Formatos: {estadisticas['formatos_diferentes']}"
        )
        self.main_window.actualizar_estadisticas(mensaje)
    
    def mostrar_vista_carnet(self):
        """Cambia a la vista de creación de carnet"""
        self.main_window.mostrar_vista_carnet()
        
        # Inicializar controlador de carnet si no está inicializado
        if self.carnet_controller is None:
            self.carnet_controller = CarnetController(
                preview_panel=self.main_window.carnet_panel.preview_panel,
                controls_panel=self.main_window.carnet_panel.controls_panel,
                employees_panel=self.main_window.carnet_panel.employees_panel
            )
        else:
            # Si ya está inicializado, refrescar la lista de empleados
            self.carnet_controller.refrescar_empleados()
    
    def exportar_data_excel(self):
        """Exporta todos los datos de la base de datos a un archivo Excel"""
        try:
            # Solicitar ruta donde guardar
            ruta_archivo, _ = QFileDialog.getSaveFileName(
                self.main_window,
                "Exportar Datos a Excel",
                f"codigos_barras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if not ruta_archivo:
                return
            
            ruta_path = Path(ruta_archivo)
            
            # Obtener formato seleccionado en el panel de generación
            formato_seleccionado = None
            if self.main_window.generation_panel:
                formato_seleccionado = self.main_window.generation_panel.combo_formato.currentText()
            
            # Exportar con el formato seleccionado para completar formatos faltantes
            exito, mensaje = self.excel_service.exportar_a_excel(ruta_path, formato_por_defecto=formato_seleccionado)
            
            if exito:
                QMessageBox.information(
                    self.main_window,
                    "Exportación Exitosa",
                    f"{mensaje}\n\nArchivo guardado en:\n{ruta_path}"
                )
                # Refrescar lista
                self.cargar_codigos()
            else:
                QMessageBox.warning(
                    self.main_window,
                    "Error al Exportar",
                    mensaje
                )
        except Exception as e:
            QMessageBox.critical(
                self.main_window,
                "Error",
                f"Error inesperado al exportar: {str(e)}"
            )
    
    def descargar_ejemplo_excel(self):
        """Genera y descarga un archivo Excel de ejemplo"""
        try:
            # Solicitar ruta donde guardar
            ruta_archivo, _ = QFileDialog.getSaveFileName(
                self.main_window,
                "Guardar Excel de Ejemplo",
                "ejemplo_importacion_codigos.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if not ruta_archivo:
                return
            
            ruta_path = Path(ruta_archivo)
            
            # Obtener formato seleccionado en el panel de generación
            formato_seleccionado = None
            if self.main_window.generation_panel:
                formato_seleccionado = self.main_window.generation_panel.combo_formato.currentText()
            
            # Generar ejemplo con el formato seleccionado
            exito, mensaje = self.excel_service.generar_excel_ejemplo(ruta_path, formato_por_defecto=formato_seleccionado)
            
            if exito:
                QMessageBox.information(
                    self.main_window,
                    "Ejemplo Generado",
                    f"{mensaje}\n\nArchivo guardado en:\n{ruta_path}\n\n"
                    "Puede abrir este archivo para ver el formato esperado."
                )
            else:
                QMessageBox.warning(
                    self.main_window,
                    "Error",
                    mensaje
                )
        except Exception as e:
            QMessageBox.critical(
                self.main_window,
                "Error",
                f"Error inesperado al generar ejemplo: {str(e)}"
            )
    
    def importar_data_excel(self):
        """Importa datos desde un archivo Excel y genera códigos de barras"""
        try:
            # Solicitar archivo a importar
            ruta_archivo, _ = QFileDialog.getOpenFileName(
                self.main_window,
                "Importar Datos desde Excel",
                "",
                "Archivos Excel (*.xlsx *.xls)"
            )
            
            if not ruta_archivo:
                return
            
            ruta_path = Path(ruta_archivo)
            
            # Crear diálogo de progreso
            progress_dialog = ProgressDialog("Importando Datos desde Excel", self.main_window)
            progress_dialog.show()
            
            # Función callback para actualizar progreso
            def actualizar_progreso(actual, total, mensaje):
                progress_dialog.actualizar_progreso(actual, total, mensaje)
            
            # Obtener formato seleccionado en el panel de generación
            formato_seleccionado = None
            if self.main_window.generation_panel:
                formato_seleccionado = self.main_window.generation_panel.combo_formato.currentText()
            
            # Importar (primera pasada: validación)
            exito, estadisticas, errores = self.excel_service.importar_desde_excel(
                ruta_path,
                callback_progreso=actualizar_progreso,
                formato_por_defecto=formato_seleccionado
            )
            
            if not exito:
                progress_dialog.close()
                QMessageBox.warning(
                    self.main_window,
                    "Error al Importar",
                    "\n".join(errores) if errores else "Error desconocido"
                )
                return
            
            # Mostrar resumen de validación
            mensaje_validacion = (
                f"Validación completada:\n\n"
                f"Total de filas: {estadisticas['total']}\n"
                f"Válidas para procesar: {estadisticas['exitosos']}\n"
                f"Duplicados: {estadisticas['duplicados']}\n"
                f"Errores: {estadisticas['errores']}\n"
                f"Validación fallida: {estadisticas['validacion_fallida']}"
            )
            
            if estadisticas['validacion_fallida'] > 0:
                respuesta = QMessageBox.question(
                    self.main_window,
                    "Códigos de Barras Inválidos Detectados",
                    f"{mensaje_validacion}\n\n"
                    "Se encontraron códigos de barras existentes que no pasaron la validación.\n"
                    "¿Desea regenerarlos automáticamente?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                regenerar_invalidos = respuesta == QMessageBox.StandardButton.Yes
            else:
                regenerar_invalidos = False
            
            if estadisticas['exitosos'] == 0:
                progress_dialog.close()
                mensaje_final = mensaje_validacion
                if errores:
                    mensaje_final += "\n\nErrores encontrados:\n" + "\n".join(errores[:10])
                    if len(errores) > 10:
                        mensaje_final += f"\n... y {len(errores) - 10} errores más"
                QMessageBox.information(self.main_window, "Sin Datos para Procesar", mensaje_final)
                return
            
            # Confirmar importación
            respuesta = QMessageBox.question(
                self.main_window,
                "Confirmar Importación",
                f"{mensaje_validacion}\n\n"
                f"¿Desea proceder a generar {estadisticas['exitosos']} código(s) de barras?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if respuesta != QMessageBox.StandardButton.Yes:
                progress_dialog.close()
                return
            
            # Segunda pasada: generar códigos
            progress_dialog.label_mensaje.setText("Generando códigos de barras...")
            progress_dialog.progress_bar.setValue(0)
            
            # Leer archivo nuevamente y procesar
            import openpyxl
            wb = openpyxl.load_workbook(str(ruta_path), data_only=True)
            ws = wb.active
            
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else "")
            
            idx_nombre = headers.index("Nombre del Empleado")
            idx_codigo_empleado = headers.index("Código de Empleado")
            # Buscar "Formato (opcional)" o "Formato" para compatibilidad
            idx_formato = None
            if "Formato (opcional)" in headers:
                idx_formato = headers.index("Formato (opcional)")
            elif "Formato" in headers:
                idx_formato = headers.index("Formato")
            
            exitosos_final = 0
            errores_final = []
            
            for row_idx in range(2, ws.max_row + 1):
                actualizar_progreso(
                    row_idx - 1,
                    ws.max_row - 1,
                    f"Generando código {row_idx - 1} de {ws.max_row - 1}..."
                )
                
                nombre_empleado = ws.cell(row=row_idx, column=idx_nombre + 1).value
                codigo_empleado = ws.cell(row=row_idx, column=idx_codigo_empleado + 1).value
                formato = ws.cell(row=row_idx, column=idx_formato + 1).value if idx_formato is not None else None
                
                if not nombre_empleado or not codigo_empleado:
                    continue
                
                nombre_empleado = str(nombre_empleado).strip()
                codigo_empleado = str(codigo_empleado).strip()
                # Usar formato del Excel si existe, sino el formato seleccionado en el dropdown, sino Code128
                formato = str(formato).strip() if formato else (formato_seleccionado or "Code128")
                
                # Validar formato - si no es Code128, cambiarlo a Code128
                formatos_validos = ["Code128", "EAN13", "EAN8", "Code39"]
                if formato not in formatos_validos:
                    formato = "Code128"  # Usar por defecto si es inválido
                elif formato != "Code128":
                    # Si el formato es válido pero no es Code128, cambiarlo a Code128
                    formato = "Code128"
                
                # Verificar si ya existe (buscar por código de empleado en descripcion)
                todos_codigos = self.db_manager.obtener_todos_codigos()
                codigo_existente = None
                for codigo in todos_codigos:
                    # codigo es: (id, codigo_barras, id_unico, fecha_creacion, nombre_empleado, descripcion, formato, nombre_archivo)
                    if len(codigo) >= 6 and codigo[5] == codigo_empleado:  # descripcion es índice 5
                        codigo_existente = codigo
                        break
                
                if codigo_existente:
                    id_db, codigo_barras, id_unico, fecha, nombre, descripcion, formato_existente, nombre_archivo = codigo_existente
                    
                    # Si hay código inválido y el usuario quiere regenerarlo
                    if regenerar_invalidos and nombre_archivo:
                        ruta_imagen = IMAGES_DIR / nombre_archivo
                        if ruta_imagen.exists():
                            valido, _ = self.barcode_service.validar_codigo_barras(ruta_imagen, id_unico)
                            if not valido:
                                # Eliminar el código existente y regenerar
                                self.db_manager.eliminar_codigo(id_db)
                                # Continuar para generar nuevo código
                            else:
                                continue  # Ya existe y es válido
                    else:
                        continue  # Ya existe, saltar
                
                # Generar código de barras
                try:
                    # Generar ID único
                    id_unico_generado = IDGenerator.generar_id_personalizado(
                        tipo="alfanumerico",
                        longitud=10,
                        incluir_nombre=False,
                        texto_personalizado=None,
                        verificar_duplicado=self.db_manager.verificar_codigo_existe
                    )
                    
                    codigo_barras, id_unico_archivo, ruta_imagen = self.barcode_service.generar_codigo_barras(
                        id_unico_generado, formato, id_unico_generado, nombre_empleado
                    )
                    
                    # Validar código generado
                    valido, mensaje_error = self.barcode_service.validar_codigo_barras(
                        ruta_imagen, id_unico_generado
                    )
                    
                    if not valido:
                        if ruta_imagen.exists():
                            ruta_imagen.unlink()
                        errores_final.append(
                            f"Fila {row_idx} ({nombre_empleado}): {mensaje_error}"
                        )
                        continue
                    
                    nombre_archivo = ruta_imagen.name
                    
                    # Guardar en base de datos
                    exito = self.db_manager.insertar_codigo(
                        codigo_barras, id_unico_generado, formato, 
                        nombre_empleado, codigo_empleado, nombre_archivo
                    )
                    
                    if not exito:
                        if ruta_imagen.exists():
                            ruta_imagen.unlink()
                        errores_final.append(
                            f"Fila {row_idx} ({nombre_empleado}): No se pudo guardar en la base de datos"
                        )
                    else:
                        exitosos_final += 1
                        
                except Exception as e:
                    errores_final.append(
                        f"Fila {row_idx} ({nombre_empleado}): {str(e)}"
                    )
            
            progress_dialog.close()
            
            # Mostrar resultado final
            mensaje_final = (
                f"Importación completada:\n\n"
                f"✓ Códigos generados exitosamente: {exitosos_final}\n"
                f"✗ Errores: {len(errores_final)}"
            )
            
            if errores_final:
                mensaje_final += "\n\nErrores encontrados:\n" + "\n".join(errores_final[:10])
                if len(errores_final) > 10:
                    mensaje_final += f"\n... y {len(errores_final) - 10} errores más"
            
            QMessageBox.information(
                self.main_window,
                "Importación Completada",
                mensaje_final
            )
            
            # Refrescar lista y estadísticas
            self.cargar_codigos()
            self.actualizar_estadisticas()
            if self.carnet_controller:
                self.carnet_controller.refrescar_empleados()
                
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al importar desde Excel: {e}", exc_info=True)
            QMessageBox.critical(
                self.main_window,
                "Error",
                f"Error inesperado al importar: {str(e)}"
            )
    
    def mostrar_vista_generacion(self):
        """Cambia a la vista de generación de códigos"""
        self.main_window.mostrar_vista_generacion()
    
    def show(self):
        """Muestra la ventana principal"""
        self.main_window.show()

