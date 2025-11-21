"""
Script para generar un archivo Excel de prueba con 2000 registros
Algunos registros tendrán formato y otros no
"""
import random
import sys
from pathlib import Path
from datetime import datetime

# Intentar importar openpyxl primero
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl no está instalado.")
    print("Asegúrate de estar usando el entorno virtual correcto.")
    print("Activa el entorno virtual con: source env/bin/activate")
    print("Luego instala con: pip install openpyxl")
    print("\nO ejecuta directamente con: ./env/bin/python3 tests/generar_excel_prueba.py")
    sys.exit(1)

# Agregar el directorio raíz al path (después de las importaciones críticas)
sys.path.insert(0, str(Path(__file__).parent.parent))

# Lista de nombres (solo nombres, sin apellidos)
NOMBRES = [
    "Juan", "María", "Carlos", "Ana", "Luis",
    "Laura", "Pedro", "Carmen", "Miguel", "Sofía",
    "Diego", "Elena", "Roberto", "Patricia", "Fernando",
    "Isabel", "Javier", "Lucía", "Ricardo", "Mónica",
    "Andrés", "Gabriela", "Sergio", "Valentina", "Daniel",
    "Camila", "Alejandro", "Natalia", "Esteban", "Andrea",
    "Felipe", "Mariana", "Gustavo", "Daniela", "Héctor",
    "Paola", "Óscar", "Carolina", "Jorge", "Adriana",
    "Manuel", "Verónica", "Raúl", "Claudia", "Eduardo",
    "Diana", "Mauricio", "Tatiana", "Cristian", "Lorena",
    "Sebastián", "Angélica", "Nicolás", "Sandra", "Fabio",
    "Jhon", "Yenny", "Camilo", "Diana Carolina", "Andrés Felipe",
    "Sara Milena", "Jorge Mario", "Lina María", "Carlos Andrés",
    "María José", "Juan David", "Laura Sofía", "Santiago", "Isabella",
    "Mateo", "Samuel", "Valeria", "David", "Alejandra",
    "Julián", "José", "Francisco", "Antonio", "Jesús",
    "Ángel", "Mario"
]

# Lista de apellidos
APELLIDOS = [
    "Pérez", "González", "Rodríguez", "Martínez", "Hernández",
    "Sánchez", "López", "García", "Torres", "Ramírez",
    "Morales", "Fernández", "Jiménez", "Ruiz", "Díaz",
    "Moreno", "Castro", "Vargas", "Ortega", "Herrera",
    "Mendoza", "Silva", "Rojas", "Medina", "Vega",
    "Paredes", "Campos", "Guzmán", "Navarro", "Fuentes",
    "Cárdenas", "Salazar", "Peña", "Ríos", "Suárez",
    "Gutiérrez", "Méndez", "Delgado", "Romero", "Aguilar",
    "Cruz", "Espinoza", "Ponce", "Núñez", "Flores",
    "Bravo", "Rivas", "Valdez", "Córdoba", "Ochoa",
    "Montoya", "Restrepo", "Zapata", "Quintero", "Arango",
    "López", "Restrepo", "Carolina", "Milena", "Mario",
    "María", "David", "Sofía"
]

FORMATOS = ["Code128", "EAN13", "EAN8", "Code39"]


def generar_registro_empleado(numero: int, incluir_formato: bool = True) -> tuple:
    """
    Genera datos de un empleado para el Excel
    
    Args:
        numero: Número secuencial del empleado
        incluir_formato: Si True, incluye formato; si False, deja el formato vacío
        
    Returns:
        Tupla con (nombres, apellidos, codigo_empleado, formato)
    """
    nombres = random.choice(NOMBRES)
    apellidos = random.choice(APELLIDOS)
    codigo_empleado = f"EMP{numero:05d}"
    
    if incluir_formato:
        formato = random.choice(FORMATOS)
    else:
        formato = None  # Sin formato
    
    return (nombres, apellidos, codigo_empleado, formato)


def generar_excel_prueba(cantidad: int = 2000, ruta_archivo: Path = None):
    """
    Genera un archivo Excel de prueba con la cantidad especificada de registros
    
    Args:
        cantidad: Número de registros a generar (por defecto 2000)
        ruta_archivo: Ruta donde guardar el archivo Excel. Si es None, se genera automáticamente
    """
    print(f"Iniciando generación de Excel con {cantidad} registros...")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos de Empleados"
    
    # Columnas (deben coincidir con lo que espera el servicio de importación)
    columnas = [
        "Nombres",              # Obligatorio
        "Apellidos",            # Obligatorio
        "Código de Empleado",   # Obligatorio
        "Formato (opcional)"    # Opcional
    ]
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    for col_idx, columna in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=columna)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Generar registros
    # Aproximadamente 60% con formato, 40% sin formato
    registros_con_formato = 0
    registros_sin_formato = 0
    
    for i in range(1, cantidad + 1):
        # Decidir si incluir formato (60% de probabilidad)
        incluir_formato = random.random() < 0.6
        
        nombres, apellidos, codigo_empleado, formato = generar_registro_empleado(i, incluir_formato)
        
        # Escribir datos en la fila
        ws.cell(row=i + 1, column=1, value=nombres)
        ws.cell(row=i + 1, column=2, value=apellidos)
        ws.cell(row=i + 1, column=3, value=codigo_empleado)
        if formato:
            ws.cell(row=i + 1, column=4, value=formato)
            registros_con_formato += 1
        else:
            registros_sin_formato += 1
        
        # Mostrar progreso cada 200 registros
        if i % 200 == 0:
            print(f"Progreso: {i}/{cantidad} registros generados...")
    
    # Ajustar ancho de columnas
    for col_idx in range(1, len(columnas) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25
    
    # Preguntar dónde guardar el archivo si no se proporciona
    if ruta_archivo is None:
        print("\n" + "="*50)
        print("SELECCIÓN DE UBICACIÓN")
        print("="*50)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_sugerido = f"excel_prueba_{cantidad}_{timestamp}.xlsx"
        ruta_sugerida = Path(__file__).parent.parent / nombre_sugerido
        
        print(f"Nombre sugerido: {nombre_sugerido}")
        print(f"Ruta sugerida: {ruta_sugerida.parent}")
        print("\nIngrese la ruta completa donde desea guardar el archivo Excel:")
        print("(Presione Enter para usar la ruta sugerida)")
        ruta_input = input("> ").strip()
        
        if ruta_input:
            ruta_archivo = Path(ruta_input)
            # Si no tiene extensión, agregarla
            if ruta_archivo.suffix.lower() != '.xlsx':
                ruta_archivo = ruta_archivo.with_suffix('.xlsx')
        else:
            ruta_archivo = ruta_sugerida
    
    # Crear directorio si no existe
    ruta_archivo.parent.mkdir(parents=True, exist_ok=True)
    
    # Guardar archivo
    print(f"\nGuardando archivo en: {ruta_archivo}...")
    wb.save(str(ruta_archivo))
    
    # Mostrar resumen
    print("\n" + "="*50)
    print("RESUMEN DE GENERACIÓN")
    print("="*50)
    print(f"Total de registros generados: {cantidad}")
    print(f"Registros con formato: {registros_con_formato} ({registros_con_formato*100/cantidad:.1f}%)")
    print(f"Registros sin formato: {registros_sin_formato} ({registros_sin_formato*100/cantidad:.1f}%)")
    print(f"Archivo guardado en: {ruta_archivo}")
    print("="*50)
    
    return ruta_archivo


if __name__ == "__main__":
    try:
        # Obtener cantidad desde argumentos de línea de comandos
        cantidad = int(sys.argv[1]) if len(sys.argv) > 1 else 2000
        
        # Obtener ruta opcional desde argumentos
        ruta_archivo = None
        if len(sys.argv) > 2:
            ruta_archivo = Path(sys.argv[2])
        
        generar_excel_prueba(cantidad, ruta_archivo)
        
    except KeyboardInterrupt:
        print("\n\nOperación cancelada por el usuario.")
        sys.exit(1)
    except ValueError as e:
        print(f"\nError: Cantidad inválida. Debe ser un número entero.")
        print(f"Uso: python generar_excel_prueba.py [cantidad] [ruta_archivo]")
        sys.exit(1)
    except Exception as e:
        print(f"\nError fatal: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

